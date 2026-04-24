#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para criar uma planilha Excel com todos os itens do MYTHARA
"""

import sqlite3
import os
from datetime import datetime
import sys
import subprocess

import subprocess
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não está instalado!")
    print("Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

PROJECT_ROOT = r"c:\Users\Wesley\Desktop\meu jogo 3d 3 os assets\🌍 MYTHARA Gurra de Raças 17-10-2025"
DB_PATH = os.path.join(PROJECT_ROOT, "Server", "gamedata.db")
EXCEL_OUTPUT = os.path.join(PROJECT_ROOT, "MYTHARA_Items_Database.xlsx")

# Cores para raridades
RARITY_COLORS = {
    "Common": "D3D3D3",      # Cinza claro
    "Uncommon": "90EE90",    # Verde claro
    "Rare": "87CEEB",        # Azul claro
    "Epic": "DDA0DD",        # Roxo
    "Legendary": "FFD700",   # Ouro
    "Divine": "FF69B4",      # Rosa
    "Mystic": "9370DB"       # Roxo médio
}

# Cores para categorias (abas)
CATEGORY_COLORS = {
    "Equipment": "0070C0",    # Azul
    "Weapons": "C00000",      # Vermelho
    "Accessories": "FFC000",  # Laranja
    "Consumables": "70AD47",  # Verde
    "Crafting Materials": "4472C4",  # Azul escuro
    "Resources": "A5A5A5",    # Cinza
    "Books": "7030A0",        # Roxo
    "Quest Items": "FF6600"   # Laranja escuro
}

def create_header_row(worksheet, row_num):
    """Cria linha de cabeçalho com formatação"""
    headers = [
        "ID", "Nome do Item", "Categoria", "Raridade", "Valor", 
        "Tipo", "Empilhável", "Máx Stack", "Descrição"
    ]
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def add_items_to_sheet(worksheet, conn, category=None):
    """Adiciona itens à planilha"""
    cursor = conn.cursor()
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Query
    if category:
        query = "SELECT * FROM Items WHERE Category = ? ORDER BY Rarity DESC, Value DESC, Name"
        cursor.execute(query, (category,))
    else:
        query = "SELECT * FROM Items ORDER BY Category, Rarity DESC, Value DESC"
        cursor.execute(query)
    
    items = cursor.fetchall()
    columns = [description[0] for description in cursor.description]
    
    row = 2
    for item_data in items:
        item = dict(zip(columns, item_data))
        
        # ID
        cell = worksheet.cell(row=row, column=1)
        cell.value = item['Id']
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Nome
        cell = worksheet.cell(row=row, column=2)
        cell.value = item['Name']
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
        cell.border = thin_border
        
        # Categoria
        cell = worksheet.cell(row=row, column=3)
        cell.value = item['Category']
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Raridade
        cell = worksheet.cell(row=row, column=4)
        cell.value = item['Rarity']
        rarity = item['Rarity']
        if rarity in RARITY_COLORS:
            cell.fill = PatternFill(start_color=RARITY_COLORS[rarity], 
                                   end_color=RARITY_COLORS[rarity], 
                                   fill_type="solid")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Valor
        cell = worksheet.cell(row=row, column=5)
        cell.value = item['Value']
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        
        # Tipo
        cell = worksheet.cell(row=row, column=6)
        cell.value = item['ItemType']
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Empilhável
        cell = worksheet.cell(row=row, column=7)
        cell.value = "Sim" if item['IsStackable'] else "Não"
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Máx Stack
        cell = worksheet.cell(row=row, column=8)
        cell.value = item['MaxStack']
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
        # Descrição
        cell = worksheet.cell(row=row, column=9)
        cell.value = item['Description']
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
        cell.border = thin_border
        
        row += 1
    
    return row - 2  # Retorna quantidade de itens adicionados

def adjust_column_widths(worksheet):
    """Ajusta largura das colunas"""
    widths = [6, 35, 18, 15, 12, 8, 12, 10, 40]
    for i, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width

def add_statistics_sheet(workbook, conn):
    """Adiciona planilha com estatísticas"""
    ws = workbook.create_sheet("📊 Estatísticas", 0)
    
    cursor = conn.cursor()
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Título
    cell = ws.cell(row=row, column=1)
    cell.value = "ESTATÍSTICAS DO BANCO DE DADOS - MYTHARA"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells(f'A{row}:B{row}')
    row += 2
    
    # Total de itens
    cursor.execute("SELECT COUNT(*) FROM Items")
    total = cursor.fetchone()[0]
    
    cell = ws.cell(row=row, column=1)
    cell.value = "Total de Itens:"
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    
    cell = ws.cell(row=row, column=2)
    cell.value = total
    cell.number_format = '#,##0'
    cell.font = Font(size=11)
    cell.border = thin_border
    row += 1
    
    # Valor total
    cursor.execute("SELECT SUM(Value) FROM Items")
    total_value = cursor.fetchone()[0] or 0
    
    cell = ws.cell(row=row, column=1)
    cell.value = "Valor Total:"
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    
    cell = ws.cell(row=row, column=2)
    cell.value = total_value
    cell.number_format = '#,##0'
    cell.font = Font(size=11)
    cell.border = thin_border
    row += 2
    
    # Itens por categoria
    ws.cell(row=row, column=1).value = "ITENS POR CATEGORIA"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    cursor.execute("SELECT Category, COUNT(*) FROM Items GROUP BY Category ORDER BY COUNT(*) DESC")
    for category, count in cursor.fetchall():
        cell = ws.cell(row=row, column=1)
        cell.value = category
        cell.border = thin_border
        
        cell = ws.cell(row=row, column=2)
        cell.value = count
        cell.number_format = '#,##0'
        cell.border = thin_border
        row += 1
    
    row += 1
    
    # Itens por raridade
    ws.cell(row=row, column=1).value = "ITENS POR RARIDADE"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    rarity_order = ["Legendary", "Divine", "Mystic", "Epic", "Rare", "Uncommon", "Common"]
    for rarity in rarity_order:
        cursor.execute("SELECT COUNT(*) FROM Items WHERE Rarity = ?", (rarity,))
        count = cursor.fetchone()[0]
        
        cell = ws.cell(row=row, column=1)
        cell.value = rarity
        cell.border = thin_border
        if rarity in RARITY_COLORS:
            cell.fill = PatternFill(start_color=RARITY_COLORS[rarity], 
                                   end_color=RARITY_COLORS[rarity], 
                                   fill_type="solid")
        
        cell = ws.cell(row=row, column=2)
        cell.value = count
        cell.number_format = '#,##0'
        cell.border = thin_border
        row += 1
    
    # Ajusta colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

def main():
    """Função principal"""
    print("\n" + "=" * 70)
    print("📊 CRIADOR DE PLANILHA EXCEL - MYTHARA ITEMS DATABASE".center(70, "="))
    print("=" * 70)
    
    try:
        if not os.path.exists(DB_PATH):
            print(f"\n❌ Banco de dados não encontrado: {DB_PATH}")
            return 1
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Cria workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove planilha padrão
        
        print(f"\n✓ Conectado ao banco de dados")
        print(f"✓ Criando planilha Excel...")
        
        # Adiciona abas por categoria
        cursor.execute("SELECT DISTINCT Category FROM Items ORDER BY Category")
        categories = [row[0] for row in cursor.fetchall()]
        
        total_items = 0
        
        for category in categories:
            print(f"  → Adicionando aba: {category}...", end="", flush=True)
            
            ws = wb.create_sheet(title=category)
            
            # Cor da aba
            if category in CATEGORY_COLORS:
                ws.sheet_properties.tabColor = CATEGORY_COLORS[category]
            
            create_header_row(ws, 1)
            items_count = add_items_to_sheet(ws, conn, category)
            adjust_column_widths(ws)
            
            # Congela a primeira linha
            ws.freeze_panes = "A2"
            
            total_items += items_count
            print(f" ✓ ({items_count} itens)")
        
        # Adiciona abas adicionais
        print(f"  → Adicionando aba: Todos os Itens...", end="", flush=True)
        ws_all = wb.create_sheet(title="📋 Todos os Itens")
        ws_all.sheet_properties.tabColor = "808080"
        create_header_row(ws_all, 1)
        add_items_to_sheet(ws_all, conn)
        adjust_column_widths(ws_all)
        ws_all.freeze_panes = "A2"
        print(" ✓")
        
        # Adiciona planilha de estatísticas
        print(f"  → Adicionando aba: Estatísticas...", end="", flush=True)
        add_statistics_sheet(wb, conn)
        print(" ✓")
        
        # Salva arquivo
        print(f"\n✓ Salvando arquivo Excel...", end="", flush=True)
        wb.save(EXCEL_OUTPUT)
        print(" ✓")
        
        # Estatísticas finais
        file_size = os.path.getsize(EXCEL_OUTPUT) / (1024 * 1024)  # MB
        
        print("\n" + "=" * 70)
        print("✅ SUCESSO! PLANILHA CRIADA COM SUCESSO".center(70, "="))
        print("=" * 70)
        
        print(f"\n📊 Resumo:")
        print(f"   • Total de itens: {total_items}")
        print(f"   • Categorias: {len(categories)}")
        print(f"   • Abas criadas: {len(categories) + 2}")
        print(f"   • Tamanho do arquivo: {file_size:.2f} MB")
        
        print(f"\n📁 Localização:")
        print(f"   {EXCEL_OUTPUT}")
        
        print(f"\n🗂️ Abas da planilha:")
        print(f"   1. 📊 Estatísticas (resumo geral)")
        for idx, category in enumerate(categories, 2):
            cursor.execute("SELECT COUNT(*) FROM Items WHERE Category = ?", (category,))
            count = cursor.fetchone()[0]
            print(f"   {idx}. {category} ({count} itens)")
        print(f"   {len(categories) + 2}. 📋 Todos os Itens ({total_items} itens)")
        
        print("\n" + "=" * 70)
        
        conn.close()
        return 0
        
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit(main())
