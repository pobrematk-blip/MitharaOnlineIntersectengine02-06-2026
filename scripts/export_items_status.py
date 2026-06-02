#!/usr/bin/env python3
import sqlite3
import os
import json
import sys
import subprocess
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Mapeamento de stats
STATS_NAMES = ["Força", "Inteligência", "Sabedoria", "Destreza", "Constituição"]
VITALS_NAMES = ["Vida", "Mana"]

# Caminhos
db_path = os.path.join(os.path.dirname(__file__), '..', 'Server', 'resources', 'gamedata.db')
output_path = os.path.join(os.path.dirname(__file__), 'MYTHARA_Items_Status_Detalhado.xlsx')

print(f"🔥 Gerando relatório de ITENS COM TODOS OS STATUS...\n")

# Conectar ao banco
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Items com Status"

# Estilos
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=9)
stat_header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
alt_fill = PatternFill(start_color="E8F0F5", end_color="E8F0F5", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers principais
headers = [
    "Nome", "Id", "Tipo", "Preço", "Raridade", "Rarity Level",
    # Dano e Defesa
    "Dano", "Tipo Dano", "Defesa", "Absorção Bloqueio", "Chance Bloqueio", "Velocidade Ataque", "Speed",
    # Crítico
    "Chance Crítico %", "Multiplicador Crítico",
    # Stats diretos
    "Força +", "Inteligência +", "Sabedoria +", "Destreza +", "Constituição +",
    # Stats Percentuais
    "Força +%", "Inteligência +%", "Sabedoria +%", "Destreza +%", "Constituição +%",
    # Vitals
    "Vida +", "Mana +",
    # Vitals Percentuais
    "Vida +%", "Mana +%",
    # Regen
    "Vida Regen", "Mana Regen",
    # Controle
    "Pode Vender", "Pode Trocar", "Pode Banco", "Pode Bag", "Stackable", "Max Stack Inv", "Max Stack Banco",
    # Misc
    "Descrição"
]

# Escrever headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Buscar todos os itens
cursor.execute("SELECT * FROM Items ORDER BY Name")
items = cursor.fetchall()

print(f"📊 Processando {len(items)} itens...\n")

row_idx = 2
for item in items:
    # Extrair dados básicos
    name = item['Name']
    item_id = item['Id']
    item_type = item['ItemType']
    price = item['Price'] or 0
    rarity = item['Rarity'] or 0
    
    # Dano e Defesa
    damage = item['Damage'] or 0
    damage_type = item['DamageType'] or 0
    defense = 0  # Items não têm defense direto, apenas defesa via armadura
    block_absorption = item['BlockAbsorption'] or 0
    block_chance = item['BlockChance'] or 0
    attack_speed_value = item['AttackSpeedValue'] or 0
    speed = item['Speed'] or 0
    
    # Crítico
    crit_chance = item['CritChance'] or 0
    crit_multiplier = item['CritMultiplier'] or 1.0
    
    # Parse JSON dos stats/vitals
    try:
        stats_given = json.loads(item['StatsGiven']) if item['StatsGiven'] else [0, 0, 0, 0, 0]
    except:
        stats_given = [0, 0, 0, 0, 0]
    
    try:
        percentage_stats = json.loads(item['PercentageStatsGiven']) if item['PercentageStatsGiven'] else [0, 0, 0, 0, 0]
    except:
        percentage_stats = [0, 0, 0, 0, 0]
    
    try:
        vitals_given = json.loads(item['VitalsGiven']) if item['VitalsGiven'] else [0, 0]
    except:
        vitals_given = [0, 0]
    
    try:
        percentage_vitals = json.loads(item['PercentageVitalsGiven']) if item['PercentageVitalsGiven'] else [0, 0]
    except:
        percentage_vitals = [0, 0]
    
    try:
        vitals_regen = json.loads(item['VitalsRegen']) if item['VitalsRegen'] else [0, 0]
    except:
        vitals_regen = [0, 0]
    
    # Controles
    can_sell = 1 if item['CanSell'] else 0
    can_trade = 1 if item['CanTrade'] else 0
    can_bank = 1 if item['CanBank'] else 0
    can_bag = 1 if item['CanBag'] else 0
    stackable = 1 if item['Stackable'] else 0
    max_inv_stack = item['MaxInventoryStack'] or 1
    max_bank_stack = item['MaxBankStack'] or 1
    
    description = item['Description'] or ""
    if len(description) > 500:
        description = description[:500] + "..."
    
    # Preparar linha
    row_data = [
        name, item_id, item_type, price, rarity,
        "⭐" * rarity if rarity <= 5 else f"Nível {rarity}",
        damage, damage_type, defense, block_absorption, block_chance, attack_speed_value, speed,
        crit_chance, crit_multiplier,
        # Stats
        stats_given[0], stats_given[1], stats_given[2], stats_given[3], stats_given[4],
        # Stats %
        percentage_stats[0], percentage_stats[1], percentage_stats[2], percentage_stats[3], percentage_stats[4],
        # Vitals
        vitals_given[0], vitals_given[1],
        # Vitals %
        percentage_vitals[0], percentage_vitals[1],
        # Regen
        vitals_regen[0], vitals_regen[1],
        # Controles
        can_sell, can_trade, can_bank, can_bag, stackable, max_inv_stack, max_bank_stack,
        # Descrição
        description
    ]
    
    # Escrever dados
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col_idx <= 14 or col_idx >= 27 else "left", 
                                   vertical="top", wrap_text=True)
        
        # Alternância de cores
        if row_idx % 2 == 0:
            cell.fill = alt_fill
    
    row_idx += 1

# Ajustar largura das colunas
column_widths = {
    1: 30,   # Nome
    2: 20,   # ID
    3: 8,    # Tipo
    4: 10,   # Preço
    5: 10,   # Raridade
    6: 12,   # Rarity visual
    7: 8,    # Dano
    8: 12,   # Tipo Dano
    9: 8,    # Defesa
    10: 10,  # Bloqueio
    11: 12,  # Chance Bloqueio
    12: 14,  # Vel Ataque
    13: 8,   # Speed
    14: 14,  # Crit Chance
    15: 14,  # Crit Mult
    # Stats e vitals - tamanho médio
    16: 10, 17: 10, 18: 10, 19: 10, 20: 10,
    21: 10, 22: 10, 23: 10, 24: 10, 25: 10,
    26: 10, 27: 10,
    28: 10, 29: 10,
    30: 10, 31: 10,
    # Controles
    32: 8, 33: 8, 34: 8, 35: 8, 36: 8, 37: 12, 38: 12,
    # Descrição
    39: 60
}

for col, width in column_widths.items():
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

# Congelar linhas
ws.freeze_panes = "A2"

# Salvar
conn.close()
wb.save(output_path)

print(f"✅ SUCESSO!")
print(f"📊 {len(items)} itens exportados com TODOS os status")
print(f"📁 Arquivo: {output_path}")
print(f"📅 Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
