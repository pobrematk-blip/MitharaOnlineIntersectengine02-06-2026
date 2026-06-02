#!/usr/bin/env python3
import sqlite3
import json
import os
import sys
import subprocess
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Mappers de enums
DAMAGE_TYPES = {
    0: "Físico",
    1: "Mágico",
    2: "Verdadeiro"
}

SPELL_TYPES = {
    0: "Ofensivo",
    1: "Defensivo",
    2: "Suporte",
    3: "Passiva"
}

COMBAT_EFFECTS = {
    0: "Nenhum",
    1: "Stun",
    2: "Silence",
    3: "Blind",
    4: "Bleed",
    5: "Poison",
    6: "Freeze",
    7: "Sleep",
    8: "Confuse"
}

TARGET_TYPES = {
    0: "Nenhum",
    1: "Inimigo",
    2: "Aliado",
    3: "Auto",
    4: "Área"
}

STATS_NAMES = ["Força", "Inteligência", "Sabedoria", "Destreza", "Constituição"]
VITALS_NAMES = ["Vida", "Mana/Stamina"]

# Caminhos
db_path = os.path.join(os.path.dirname(__file__), '..', 'Server', 'resources', 'gamedata.db')
output_path = os.path.join(os.path.dirname(__file__), 'MYTHARA_Skills_Detalhado.xlsx')

print(f"🔥 Gerando relatório de SKILLS COM TODOS OS EFEITOS E CLASSES...\n")

# Conectar ao banco
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()

# Criar workbook
wb = Workbook()

# ================= SHEET 1: SKILLS DETALHADAS =================
ws_skills = wb.active
ws_skills.title = "Skills Detalhadas"

# Estilos
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=9)
alt_fill = PatternFill(start_color="E8F0F5", end_color="E8F0F5", fill_type="solid")
damage_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers para skills
skill_headers = [
    "Nome", "ID", "Tipo", "Classe/Folder",
    # Combat
    "Tipo Dano", "Alcance", "Raio Hit", "Crítico %", "Crit Mult", "Duração",
    "Efeito", "Alvo", "Escala Dano", "Stat Escala",
    # Custos
    "Vida Cost", "Mana/Stamina Cost", "Cooldown (ms)", "Cooldown Grupo",
    # Dados
    "Força Dado", "Int Dado", "Sab Dado", "Dex Dado", "Con Dado",
    "Força % Dado", "Int % Dado", "Sab % Dado", "Dex % Dado", "Con % Dado",
    "Vida Dado", "Mana Dado", "Vida Regen", "Mana Regen",
    # Outros
    "HoT/DoT", "Interval DoT", "Bound", "Descrição"
]

for col_idx, header in enumerate(skill_headers, 1):
    cell = ws_skills.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Buscar todas as skills
cursor.execute("SELECT * FROM Spells ORDER BY Name")
spells = cursor.fetchall()

print(f"📊 Processando {len(spells)} skills...\n")

row_idx = 2
for spell in spells:
    name = spell['Name']
    spell_id = spell['Id']
    spell_type = SPELL_TYPES.get(spell['SpellType'] or 0, "Desconhecido")
    folder = spell['Folder'] or "Geral"
    
    # Combat
    damage_type = DAMAGE_TYPES.get(spell['Combat_DamageType'] or 0, "Desconhecido")
    cast_range = spell['Combat_CastRange'] or 0
    hit_radius = spell['Combat_HitRadius'] or 0
    crit_chance = spell['Combat_CritChance'] or 0
    crit_mult = spell['Combat_CritMultiplier'] or 1.0
    duration = spell['Combat_Duration'] or 0
    
    effect = COMBAT_EFFECTS.get(spell['Combat_Effect'] or 0, "Nenhum")
    target_type = TARGET_TYPES.get(spell['Combat_TargetType'] or 0, "Desconhecido")
    scaling = spell['Combat_Scaling'] or 0
    scaling_stat = spell['Combat_ScalingStat'] or 0
    
    # Custos
    try:
        vital_cost = json.loads(spell['VitalCost']) if spell['VitalCost'] else [0, 0]
    except:
        vital_cost = [0, 0]
    
    life_cost = vital_cost[0] if len(vital_cost) > 0 else 0
    mana_cost = vital_cost[1] if len(vital_cost) > 1 else 0
    
    cooldown = spell['CooldownDuration'] or 0
    cooldown_group = spell['CooldownGroup'] or ""
    
    # Dados de Stats
    try:
        stat_diff = json.loads(spell['StatDiff']) if spell['StatDiff'] else [0, 0, 0, 0, 0]
    except:
        stat_diff = [0, 0, 0, 0, 0]
    
    try:
        percentage_stat_diff = json.loads(spell['PercentageStatDiff']) if spell['PercentageStatDiff'] else [0, 0, 0, 0, 0]
    except:
        percentage_stat_diff = [0, 0, 0, 0, 0]
    
    try:
        vital_diff = json.loads(spell['VitalDiff']) if spell['VitalDiff'] else [0, 0]
    except:
        vital_diff = [0, 0]
    
    # Outros
    hotdot = "Sim" if spell['Combat_HoTDoT'] else "Não"
    hotdot_interval = spell['Combat_HotDotInterval'] or 0
    bound = "Sim" if spell['Bound'] else "Não"
    
    description = spell['Description'] or ""
    if len(description) > 500:
        description = description[:500] + "..."
    
    # Preparar linha
    row_data = [
        name, spell_id, spell_type, folder,
        damage_type, cast_range, hit_radius, crit_chance, crit_mult, duration,
        effect, target_type, scaling, scaling_stat,
        life_cost, mana_cost, cooldown, cooldown_group,
        stat_diff[0], stat_diff[1], stat_diff[2], stat_diff[3], stat_diff[4],
        percentage_stat_diff[0], percentage_stat_diff[1], percentage_stat_diff[2], percentage_stat_diff[3], percentage_stat_diff[4],
        vital_diff[0], vital_diff[1], vital_diff[0], vital_diff[1],  # Regen
        hotdot, hotdot_interval, bound,
        description
    ]
    
    # Escrever dados
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_skills.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col_idx <= 14 or col_idx >= 31 else "left",
            vertical="top",
            wrap_text=True
        )
        
        if row_idx % 2 == 0:
            cell.fill = alt_fill
    
    row_idx += 1

# Ajustar colunas
col_widths = {
    1: 25, 2: 20, 3: 12, 4: 20,
    5: 12, 6: 8, 7: 8, 8: 10, 9: 10, 10: 10,
    11: 14, 12: 12, 13: 12, 14: 12,
    15: 10, 16: 14, 17: 12, 18: 15,
    19: 10, 20: 10, 21: 10, 22: 10, 23: 10,
    24: 10, 25: 10, 26: 10, 27: 10, 28: 10,
    29: 10, 30: 10, 31: 12, 32: 12,
    33: 8, 34: 8, 35: 8, 36: 60
}

for col, width in col_widths.items():
    ws_skills.column_dimensions[ws_skills.cell(row=1, column=col).column_letter].width = width

ws_skills.freeze_panes = "A2"

# ================= SHEET 2: SKILLS POR CLASSE =================
ws_class_skills = wb.create_sheet("Skills por Classe")

class_headers = ["Classe", "Skill", "Level Obtém", "Tipo", "Tipo Dano", "Cooldown", "Mana Cost"]

for col_idx, header in enumerate(class_headers, 1):
    cell = ws_class_skills.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Buscar todas as classes e suas skills
cursor.execute("SELECT Name, Id, Spells FROM Classes ORDER BY Name")
classes = cursor.fetchall()

# Criar mapa de spell_id -> spell_info
spell_map = {}
for spell in spells:
    spell_map[spell['Id']] = {
        'name': spell['Name'],
        'type': SPELL_TYPES.get(spell['SpellType'] or 0, ""),
        'damage_type': DAMAGE_TYPES.get(spell['Combat_DamageType'] or 0, ""),
        'cooldown': spell['CooldownDuration'] or 0,
        'mana_cost': 0
    }
    # Extrair custo de mana
    try:
        vital_cost = json.loads(spell['VitalCost']) if spell['VitalCost'] else [0, 0]
        spell_map[spell['Id']]['mana_cost'] = vital_cost[1] if len(vital_cost) > 1 else 0
    except:
        pass

row_idx = 2
for cls in classes:
    class_name = cls['Name']
    try:
        class_spells = json.loads(cls['Spells']) if cls['Spells'] else []
    except:
        class_spells = []
    
    for spell_info in class_spells:
        spell_id = spell_info.get('Id')
        level = spell_info.get('Level', 1)
        
        if spell_id in spell_map:
            spell_data = spell_map[spell_id]
            row_data = [
                class_name,
                spell_data['name'],
                level,
                spell_data['type'],
                spell_data['damage_type'],
                spell_data['cooldown'],
                spell_data['mana_cost']
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_class_skills.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col_idx in [3, 6, 7] else "left", vertical="top", wrap_text=True)
                
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
            
            row_idx += 1

# Ajustar colunas
ws_class_skills.column_dimensions['A'].width = 25
ws_class_skills.column_dimensions['B'].width = 25
ws_class_skills.column_dimensions['C'].width = 12
ws_class_skills.column_dimensions['D'].width = 12
ws_class_skills.column_dimensions['E'].width = 12
ws_class_skills.column_dimensions['F'].width = 12
ws_class_skills.column_dimensions['G'].width = 12

ws_class_skills.freeze_panes = "A2"

# Salvar
conn.close()
wb.save(output_path)

print(f"✅ SUCESSO!")
print(f"📊 {len(spells)} skills exportadas com TODOS os detalhes")
print(f"📁 Arquivo: {output_path}")
print(f"📅 Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
