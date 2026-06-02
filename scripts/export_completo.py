#!/usr/bin/env python3
import sqlite3
import os
import sys
import subprocess
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Caminhos
db_path = os.path.join(os.path.dirname(__file__), '..', 'Server', 'resources', 'gamedata.db')
output_path = os.path.join(os.path.dirname(__file__), 'MYTHARA_Database_Completo.xlsx')

print(f"🔥 Gerando relatório Excel COMPLETO com todos os detalhes...\n")

# Conectar ao banco
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# Estilos
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
alt_fill = PatternFill(start_color="E8F0F5", end_color="E8F0F5", fill_type="solid")

def create_detailed_sheet(table_name, limit=None):
    """Cria uma sheet com TODOS os dados da tabela"""
    if limit:
        query = f"SELECT * FROM {table_name} LIMIT {limit}"
    else:
        query = f"SELECT * FROM {table_name}"
    
    cursor.execute(query)
    rows = cursor.fetchall()
    
    if not rows:
        print(f"⚠️  {table_name}: sem dados")
        return
    
    # Limitar nome da sheet (Excel tem limite de 31 caracteres)
    sheet_name = table_name[:31]
    ws = wb.create_sheet(sheet_name)
    
    # Cabeçalhos
    columns = [description[0] for description in cursor.description]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Dados
    for row_idx, row in enumerate(rows, 2):
        # Alternar cores de linhas
        row_fill = alt_fill if row_idx % 2 == 0 else None
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            try:
                # Limpar valores muito grandes (manter JSON intacto mas truncar display)
                if isinstance(value, bytes):
                    cell.value = "[BLOB - Dados Binários]"
                elif isinstance(value, str) and len(value) > 1000:
                    cell.value = value[:1000] + "...[truncado]"
                else:
                    cell.value = value
            except (UnicodeDecodeError, TypeError):
                cell.value = "[Dados Não-Legíveis]"
            
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if row_fill:
                cell.fill = row_fill
    
    # Ajustar largura das colunas
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Calcular largura baseado no conteúdo
        max_length = len(str(col_name))
        for row in rows[:100]:  # Verificar primeiras 100 linhas
            if col_idx <= len(row) and row[col_idx-1]:
                max_length = max(max_length, len(str(row[col_idx-1])[:100]))
        width = min(max_length + 2, 80)  # Máximo 80
        ws.column_dimensions[col_letter].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    print(f"✅ {table_name}: {len(rows)} registros com {len(columns)} campos")

# ================= CRIAR SHEETS DETALHADAS =================
print("📋 Criando abas detalhadas...\n")

# Dados principais
create_detailed_sheet("Items")           # Todos os 1177+ itens
create_detailed_sheet("Spells")          # Todas as 212 magias
create_detailed_sheet("Npcs")            # Todos os 121 NPCs
create_detailed_sheet("Classes")         # Todas as 109 classes
create_detailed_sheet("Quests")          # Todas as 113 missões
create_detailed_sheet("Maps")            # Todos os 525 mapas
create_detailed_sheet("Events")          # Todos os 1687 eventos
create_detailed_sheet("Crafts")          # Todas as 384 receitas
create_detailed_sheet("CraftingTables")  # Todas as 51 mesas
create_detailed_sheet("Animations")      # Todas as 136 animações
create_detailed_sheet("Resources")       # Todos os 46 recursos
create_detailed_sheet("Projectiles")     # Todos os 26 projéteis
create_detailed_sheet("Shops")           # Todas as 22 lojas
create_detailed_sheet("Pets")            # Todos os 4 pets
create_detailed_sheet("Tilesets")        # Todos os 279 tilesets

# Dados de configuração
create_detailed_sheet("Time")
create_detailed_sheet("UserVariables")
create_detailed_sheet("PlayerVariables")
create_detailed_sheet("ServerVariables")
create_detailed_sheet("GuildVariables")

print("\n" + "="*60)

# ================= CRIAR SHEET DE RESUMO DETALHADO =================
print("\n📊 Criando sheet de resumo...\n")

ws_summary = wb.create_sheet("RESUMO", 0)

# Contar todas as tabelas
stats = []
cursor.execute("SELECT COUNT(*) FROM Items")
stats.append(["Items (Itens/Equipamentos)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Spells")
stats.append(["Spells (Magias/Habilidades)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Npcs")
stats.append(["NPCs (Personagens)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Classes")
stats.append(["Classes", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Quests")
stats.append(["Quests (Missões)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Maps")
stats.append(["Maps (Mapas do Mundo)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Events")
stats.append(["Events (Eventos)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Crafts")
stats.append(["Crafts (Receitas)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM CraftingTables")
stats.append(["Crafting Tables (Mesas)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Animations")
stats.append(["Animations (Animações)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Resources")
stats.append(["Resources (Recursos)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Projectiles")
stats.append(["Projectiles (Projéteis)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Shops")
stats.append(["Shops (Lojas NPC)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Pets")
stats.append(["Pets (Bichos de Estimação)", cursor.fetchone()[0]])

cursor.execute("SELECT COUNT(*) FROM Tilesets")
stats.append(["Tilesets (Conjuntos)", cursor.fetchone()[0]])

# Escrever no resumo
ws_summary.cell(row=1, column=1).value = "TABELA"
ws_summary.cell(row=1, column=2).value = "QUANTIDADE"

for idx, cell in enumerate([ws_summary.cell(row=1, column=1), ws_summary.cell(row=1, column=2)], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, (table_name, count) in enumerate(stats, 2):
    ws_summary.cell(row=row_idx, column=1).value = table_name
    ws_summary.cell(row=row_idx, column=2).value = count
    
    for col in [1, 2]:
        cell = ws_summary.cell(row=row_idx, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if row_idx % 2 == 0:
            cell.fill = alt_fill

ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 20

# ================= SALVAR =================
conn.close()
wb.save(output_path)

total_tabelas = len(stats)
total_registros = sum([count for _, count in stats])

print("\n" + "="*60)
print(f"✅ SUCESSO!")
print(f"📊 Total de {total_tabelas} tabelas exportadas")
print(f"📋 Total de {total_registros:,} registros")
print(f"\n📁 Arquivo: {output_path}")
print(f"📅 Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
print("="*60)
