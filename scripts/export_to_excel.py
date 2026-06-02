#!/usr/bin/env python3
import sqlite3
import os
from datetime import datetime
import sys
import subprocess

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ Falta: pip install openpyxl")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Caminhos
db_path = os.path.join(os.path.dirname(__file__), '..', 'Server', 'resources', 'gamedata.db')
output_path = os.path.join(os.path.dirname(__file__), 'MYTHARA_Database_Report.xlsx')

print(f"📊 Gerando relatório Excel detalhado...\n")

# Conectar ao banco
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()

# Criar workbook
wb = Workbook()
wb.remove(wb.active)  # Remove sheet padrão

# Estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet_from_table(table_name, query=None):
    """Cria uma sheet no Excel com os dados da tabela"""
    if query is None:
        query = f"SELECT * FROM {table_name} LIMIT 1000"
    
    cursor.execute(query)
    rows = cursor.fetchall()
    
    if not rows:
        return
    
    ws = wb.create_sheet(table_name[:31])  # Limite de 31 chars no Excel
    
    # Cabeçalhos
    columns = [description[0] for description in cursor.description]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Dados
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Limpar valores muito grandes
            if isinstance(value, str) and len(value) > 500:
                cell.value = value[:500] + "..."
            else:
                cell.value = value
            
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Ajustar largura das colunas
    for col_idx, col_name in enumerate(columns, 1):
        max_length = max(
            len(str(col_name)),
            max([len(str(row[col_idx-1])) if row[col_idx-1] else 0 for row in rows], default=0)
        )
        width = min(max_length + 2, 60)  # Máximo 60
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    print(f"✅ {table_name}: {len(rows)} registros")

# Criar sheets principais
print("📋 Gerando sheets...\n")

create_sheet_from_table("Items", "SELECT Name, Id, ItemType, Price, Stackable, CanTrade, CanSell, Description FROM Items LIMIT 1000")
create_sheet_from_table("Spells", "SELECT Name, Id, SpellType, Combat_DamageType, Combat_CastRange, Description FROM Spells LIMIT 500")
create_sheet_from_table("Npcs", "SELECT Name, Id, Level, Experience, Damage, SightRange, SpellFrequency FROM Npcs LIMIT 500")
create_sheet_from_table("Quests", "SELECT Name, Id, Repeatable, OrderValue, Quitable FROM Quests LIMIT 300")
create_sheet_from_table("Classes", "SELECT Name, Id, BaseExp, BasePoints FROM Classes LIMIT 200")
create_sheet_from_table("Maps", "SELECT Name, Id, IsIndoors, Brightness, Music FROM Maps LIMIT 1000")
create_sheet_from_table("Events", "SELECT Name, Id, CommonEvent, Global FROM Events LIMIT 500")
create_sheet_from_table("Crafts", "SELECT Name, Id, ItemId, Quantity, Time, FailureChance FROM Crafts LIMIT 500")
create_sheet_from_table("Resources", "SELECT Name, Id, MinHp, MaxHp, SpawnDuration FROM Resources")
create_sheet_from_table("Animations", "SELECT Name, Id, Sound, CompleteSound FROM Animations")
create_sheet_from_table("Shops", "SELECT Name, Id, DefaultCurrency FROM Shops")

# Sheet de resumo
print("📊 Gerando sheet de resumo...\n")

ws_summary = wb.create_sheet("RESUMO GERAL", 0)

# Contar tabelas
cursor.execute("SELECT COUNT(*) FROM Items")
items_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Spells")
spells_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Npcs")
npcs_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Quests")
quests_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Maps")
maps_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Classes")
classes_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Events")
events_count = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM Crafts")
crafts_count = cursor.fetchone()[0]

# Dados do resumo
summary_data = [
    ["TABELA", "QUANTIDADE"],
    ["Items (Itens)", items_count],
    ["Spells (Magias)", spells_count],
    ["Npcs (Personagens)", npcs_count],
    ["Quests (Missões)", quests_count],
    ["Maps (Mapas)", maps_count],
    ["Classes", classes_count],
    ["Events (Eventos)", events_count],
    ["Crafts (Receitas)", crafts_count],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if row_idx == 1:  # Cabeçalho
            cell.fill = header_fill
            cell.font = header_font

ws_summary.column_dimensions['A'].width = 30
ws_summary.column_dimensions['B'].width = 20

# Salvar
conn.close()
wb.save(output_path)

print(f"\n✅ Arquivo salvo em: {output_path}")
print(f"📁 Pasta: {os.path.dirname(output_path)}")
print(f"📅 Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
